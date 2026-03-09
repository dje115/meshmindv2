"""XLSX and XLS extraction tests."""

from __future__ import annotations

import tempfile
from pathlib import Path

from meshmind_worker_docproc.extractors import EXTRACTORS
from meshmind_worker_docproc.models import ExtractionStatus

PROVENANCE = {"extension": "xlsx"}


def test_xlsx_extraction() -> None:
    """XLSX with sheets produces sheet-aware output."""
    import openpyxl

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        p = Path(f.name)
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "alpha"
        ws["B2"] = 1
        wb.save(p)
        result = EXTRACTORS["xlsx"](p, PROVENANCE)
        assert result.success
        assert result.document.extraction_metadata.status == ExtractionStatus.SUCCESS
        assert result.document.extraction_metadata.sheet_count == 1
        assert len(result.document.sheets) == 1
        assert result.document.sheets[0].sheet_name == "Data"
        assert "alpha" in result.document.full_text
        assert "Name" in result.document.full_text
    finally:
        p.unlink(missing_ok=True)
